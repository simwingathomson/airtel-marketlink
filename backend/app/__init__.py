from datetime import datetime, date, timedelta
import csv
import io
import os
import tempfile
from functools import wraps
from flask import Flask, Response, jsonify, redirect, render_template, request, g, session, url_for, send_file
from flask_cors import CORS
from werkzeug.exceptions import BadRequest, Forbidden, NotFound, Unauthorized
from werkzeug.utils import secure_filename
from sqlalchemy import or_, func
from openpyxl import Workbook, load_workbook
try:
    from flask_limiter import Limiter
    from flask_limiter.util import get_remote_address
except Exception:
    Limiter = None
    get_remote_address = None
from .extensions import db
from .models import *

def database_uri():
    uri = os.getenv("DATABASE_URL", "").strip()
    if not uri:
        return "sqlite:///marketlink.db"
    if uri.startswith("postgres://"):
        uri = "postgresql://" + uri[len("postgres://"):]
    if uri.startswith("postgresql://") and "+" not in uri.split("://", 1)[0]:
        uri = "postgresql+psycopg://" + uri[len("postgresql://"):]
    return uri

def create_app(test_config=None):
    app = Flask(__name__)
    app.config.update(
        SECRET_KEY=os.getenv("SECRET_KEY", "dev-change-me"),
        SQLALCHEMY_DATABASE_URI=database_uri(),
        SQLALCHEMY_ENGINE_OPTIONS={"pool_pre_ping": True, "pool_recycle": 300},
        SQLALCHEMY_TRACK_MODIFICATIONS=False,
        MAX_CONTENT_LENGTH=8 * 1024 * 1024,
        UPLOAD_FOLDER="uploads",
    )
    if test_config:
        app.config.update(test_config)
    CORS(app)
    db.init_app(app)
    limiter = Limiter(get_remote_address, app=app, default_limits=["200 per hour"]) if Limiter else None

    @app.after_request
    def security_headers(response):
        response.headers.setdefault("X-Content-Type-Options", "nosniff")
        response.headers.setdefault("X-Frame-Options", "DENY")
        response.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
        response.headers.setdefault("Permissions-Policy", "geolocation=(), microphone=(), camera=()")
        return response

    @app.errorhandler(Exception)
    def handle_error(exc):
        code = getattr(exc, "code", 500)
        if code == 500:
            app.logger.exception(exc)
        return jsonify(error=getattr(exc, "description", "Internal server error")), code

    def serialize(obj):
        data = {}
        for c in obj.__table__.columns:
            v = getattr(obj, c.name)
            if isinstance(v, (datetime, date)):
                v = v.isoformat()
            data[c.name] = v
        return data


    def csv_response(filename, rows):
        output = io.StringIO()
        if rows:
            writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            writer.writerows(rows)
        return Response(output.getvalue(), mimetype="text/csv", headers={"Content-Disposition": f"attachment; filename={filename}"})
    def parse_date(value):
        return date.fromisoformat(value) if value else None

    def auth_required(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            header = request.headers.get("Authorization", "")
            token = header.replace("Bearer ", "", 1)
            user = User.query.filter_by(token=token, active=True).first()
            if not user or not user.token_expires_at or user.token_expires_at < datetime.utcnow():
                raise Unauthorized("Authentication required")
            g.current_user = user
            return fn(*args, **kwargs)
        return wrapper

    def roles_required(*roles):
        def outer(fn):
            @wraps(fn)
            def wrapper(*args, **kwargs):
                if g.current_user.role != "Administrator" and g.current_user.role not in roles:
                    raise Forbidden("You do not have permission for this action")
                return fn(*args, **kwargs)
            return wrapper
        return outer

    def log(action, record=None, description=""):
        db.session.add(AuditLog(user_id=getattr(g, "current_user", None) and g.current_user.id,
                               action=action, record_type=record.__class__.__name__ if record else None,
                               record_id=getattr(record, "id", None), description=description))

    def crud_routes(url, model, required=("Administrator", "ZBM", "TSM", "TL")):
        endpoint = url.strip("/").replace("/", "_")
        @app.get(url, endpoint=f"{endpoint}_list")
        @auth_required
        def list_items(model=model):
            q = model.query
            search = request.args.get("search")
            if search:
                cols = [c for c in model.__table__.columns if str(c.type).startswith("VARCHAR")]
                q = q.filter(or_(*[c.ilike(f"%{search}%") for c in cols]))
            return jsonify([serialize(x) for x in q.limit(int(request.args.get("limit", 200))).all()])
        @app.post(url, endpoint=f"{endpoint}_create")
        @auth_required
        @roles_required(*required)
        def create_item(model=model):
            item = model()
            apply_payload(item, request.get_json() or {})
            db.session.add(item); db.session.flush(); log("create", item); db.session.commit()
            return jsonify(serialize(item)), 201
        @app.get(f"{url}/<int:item_id>", endpoint=f"{endpoint}_get")
        @auth_required
        def get_item(item_id, model=model):
            return jsonify(serialize(model.query.get_or_404(item_id)))
        @app.patch(f"{url}/<int:item_id>", endpoint=f"{endpoint}_patch")
        @auth_required
        @roles_required(*required)
        def patch_item(item_id, model=model):
            item = model.query.get_or_404(item_id)
            apply_payload(item, request.get_json() or {})
            log("update", item); db.session.commit()
            return jsonify(serialize(item))
        @app.delete(f"{url}/<int:item_id>", endpoint=f"{endpoint}_delete")
        @auth_required
        @roles_required("Administrator")
        def delete_item(item_id, model=model):
            item = model.query.get_or_404(item_id)
            log("delete", item); db.session.delete(item); db.session.commit()
            return "", 204

    def apply_payload(item, payload):
        dates = {"joining_date", "registration_date", "date_assigned", "date", "required_date", "due_date", "audit_date", "last_audit", "next_audit"}
        for k, v in payload.items():
            if not hasattr(item, k): continue
            setattr(item, k, parse_date(v) if k in dates and isinstance(v, str) else v)

    def validate_resource_deployment(employee_id, market_id, booth_id, partner_id):
        required = {
            "chabeba employee": employee_id,
            "market": market_id,
            "booth": booth_id,
            "channel partner": partner_id,
        }
        missing = [name for name, value in required.items() if not value]
        if missing:
            raise BadRequest("Resource deployment requires " + ", ".join(missing))

        employee = Employee.query.get(employee_id)
        market = Market.query.get(market_id)
        booth = Booth.query.get(booth_id)
        partner = ChannelPartner.query.get(partner_id)
        if not employee:
            raise BadRequest("Chabeba employee not found")
        if employee.role != "Chabeba":
            raise BadRequest("Deployed resources must be assigned to a Chabeba employee")
        if not market:
            raise BadRequest("Market not found")
        if not booth:
            raise BadRequest("Booth not found")
        if booth.market_id != market.id:
            raise BadRequest("Booth must belong to the selected market")
        if not partner:
            raise BadRequest("Channel partner not found")
        if partner.market_id != market.id:
            raise BadRequest("Channel partner must belong to the selected market")
        if partner.booth_id and partner.booth_id != booth.id:
            raise BadRequest("Channel partner must belong to the selected booth")
        return employee, market, booth, partner


    web_modules = {
        "people": {"key":"people", "label":"People", "model":Employee, "columns":["id","employee_code","full_name","role","employment_status"], "fields":[("employee_code","Employee ID"),("full_name","Full name"),("role","Role"),("phone","Phone"),("email","Email")]},
        "districts": {"key":"districts", "label":"Districts", "model":District, "columns":["id","district_code","name","region","status"], "fields":[("district_code","District ID"),("name","District name"),("region","Region"),("status","Status")]},
        "territories": {"key":"territories", "label":"Territories", "model":Territory, "columns":["id","territory_code","name","district_id","status"], "fields":[("territory_code","Territory ID"),("name","Territory name"),("district_id","District ID"),("status","Status")]},
        "markets": {"key":"markets", "label":"Markets", "model":Market, "columns":["id","market_code","name","territory_id","status"], "fields":[("market_code","Market ID"),("name","Market name"),("district_id","District ID"),("territory_id","Territory ID"),("location","Physical location")]},
        "booths": {"key":"booths", "label":"Booths", "model":Booth, "columns":["id","booth_code","market_id","location","status"], "fields":[("booth_code","Booth ID"),("market_id","Market ID"),("location","Location"),("status","Status")]},
        "partners": {"key":"partners", "label":"Partners", "model":ChannelPartner, "columns":["id","partner_code","business_name","market_id","status"], "fields":[("partner_code","Partner ID"),("business_name","Business name"),("contact_person","Owner/contact"),("market_id","Market ID"),("phone","Phone")]},
        "resources": {"key":"resources", "label":"Resources", "model":Resource, "columns":["id","resource_code","name","status","assigned_employee_id","market_id","booth_id","partner_id"], "fields":[("resource_code","Resource ID"),("name","Name"),("category_id","Category ID"),("condition","Condition"),("status","Status")]},
        "assignments": {"key":"assignments", "label":"Assignments", "model":ResourceAssignment, "columns":["id","resource_id","employee_id","market_id","booth_id","partner_id","date_assigned"], "fields":[("resource_id","Resource ID"),("employee_id","Chabeba employee ID"),("market_id","Market ID"),("booth_id","Booth ID"),("partner_id","Channel partner ID"),("condition","Condition"),("notes","Notes")]},
        "transfers": {"key":"transfers", "label":"Transfers", "model":ResourceTransfer, "columns":["id","resource_id","from_employee_id","to_employee_id","from_market_id","to_market_id","date"], "fields":[("resource_id","Resource ID"),("to_employee_id","To Chabeba employee ID"),("to_market_id","To market ID"),("booth_id","To booth ID"),("partner_id","To channel partner ID"),("reason","Reason"),("date","Date YYYY-MM-DD")]},
        "returns": {"key":"returns", "label":"Returns", "model":ResourceReturn, "columns":["id","resource_id","employee_id","condition","date"], "fields":[("resource_id","Resource ID"),("employee_id","Employee ID"),("condition","Condition"),("reason","Reason"),("notes","Notes")]},
        "requests": {"key":"requests", "label":"Requests", "model":ResourceRequest, "columns":["id","requested_by_id","category_id","required_date","status"], "fields":[("requested_by_id","Requested by employee ID"),("category_id","Category ID"),("reason","Reason"),("market_id","Market ID"),("required_date","Required date YYYY-MM-DD")]},
        "incidents": {"key":"incidents", "label":"Incidents", "model":ResourceIncident, "columns":["id","resource_id","incident_type","status","date"], "fields":[("resource_id","Resource ID"),("incident_type","Lost/Damaged/Stolen/Faulty"),("description","Description"),("location","Location")]},
        "tasks": {"key":"tasks", "label":"Tasks", "model":FieldTask, "columns":["id","title","assigned_employee_id","priority","status"], "fields":[("title","Task title"),("description","Description"),("assigned_employee_id","Assigned employee ID"),("priority","Priority"),("due_date","Due date YYYY-MM-DD")]},
        "audits": {"key":"audits", "label":"Audits", "model":MarketAudit, "columns":["id","market_id","auditor_id","audit_date","status"], "fields":[("market_id","Market ID"),("auditor_id","Auditor employee ID"),("discrepancies","Discrepancies"),("status","Status")]},
        "products": {"key":"products", "label":"Products", "model":Product, "columns":["id","sku","name","unit_price","minimum_stock","status"], "fields":[("sku","SKU"),("name","Product name"),("category_id","Category ID"),("unit","Unit"),("unit_price","Unit price"),("minimum_stock","Minimum stock"),("status","Status")]},
        "customers": {"key":"customers", "label":"Customers", "model":Customer, "columns":["id","customer_code","name","phone","location","status"], "fields":[("customer_code","Customer ID"),("name","Name"),("phone","Phone"),("email","Email"),("location","Location"),("market_id","Market ID"),("assigned_employee_id","Assigned employee ID"),("status","Status"),("notes","Notes")]},
        "inventory_locations": {"key":"inventory_locations", "label":"Inventory Locations", "model":InventoryLocation, "columns":["id","location_code","name","location_type","status"], "fields":[("location_code","Location ID"),("name","Name"),("location_type","Type"),("market_id","Market ID"),("booth_id","Booth ID"),("partner_id","Partner ID"),("status","Status")]},
        "stock": {"key":"stock", "label":"Stock", "model":InventoryStock, "columns":["id","product_id","location_id","quantity","updated_at"], "fields":[("product_id","Product ID"),("location_id","Location ID"),("quantity","Quantity")]},
        "stock_movements": {"key":"stock_movements", "label":"Stock Movements", "model":StockMovement, "columns":["id","movement_code","product_id","movement_type","quantity","created_at"], "fields":[("movement_code","Movement ID"),("product_id","Product ID"),("from_location_id","From location ID"),("to_location_id","To location ID"),("quantity","Quantity"),("movement_type","Type"),("notes","Notes")]},
        "orders": {"key":"orders", "label":"Orders", "model":Order, "columns":["id","order_number","customer_id","partner_id","status","total_amount"], "fields":[("order_number","Order number"),("customer_id","Customer ID"),("partner_id","Partner ID"),("status","Status"),("order_date","Order date YYYY-MM-DD"),("due_date","Due date YYYY-MM-DD"),("total_amount","Total amount"),("notes","Notes")]},
        "sales": {"key":"sales", "label":"Sales", "model":Sale, "columns":["id","sale_number","customer_id","partner_id","sale_date","total_amount"], "fields":[("sale_number","Sale number"),("customer_id","Customer ID"),("partner_id","Partner ID"),("employee_id","Employee ID"),("market_id","Market ID"),("sale_date","Sale date YYYY-MM-DD"),("total_amount","Total amount"),("status","Status"),("notes","Notes")]},
        "visits": {"key":"visits", "label":"Visits", "model":Visit, "columns":["id","title","visit_date","assigned_employee_id","status","priority"], "fields":[("title","Visit title"),("partner_id","Partner ID"),("customer_id","Customer ID"),("location","Location"),("visit_date","Visit date YYYY-MM-DD"),("start_time","Start time"),("end_time","End time"),("assigned_employee_id","Assigned employee ID"),("purpose","Purpose"),("status","Status"),("priority","Priority"),("notes","Notes")]},
        "plans": {"key":"plans", "label":"Plans", "model":Plan, "columns":["id","title","frequency","start_date","status","priority"], "fields":[("title","Plan title"),("description","Description"),("start_date","Start date YYYY-MM-DD"),("end_date","End date YYYY-MM-DD"),("frequency","Frequency"),("assigned_employee_id","Assigned employee ID"),("priority","Priority"),("status","Status"),("reminder_days","Reminder days")]},
        "preferences": {"key":"preferences", "label":"Notification Preferences", "model":NotificationPreference, "columns":["id","user_id","notification_type","in_app_enabled","email_enabled"], "fields":[("user_id","User ID"),("notification_type","Type"),("in_app_enabled","In app true/false"),("email_enabled","Email true/false"),("sms_enabled","SMS true/false")]},
        "import_jobs": {"key":"import_jobs", "label":"Import Jobs", "model":ImportJob, "columns":["id","filename","data_type","status","total_rows","imported_rows"], "fields":[("filename","Filename"),("data_type","Data type"),("sheet_name","Sheet"),("status","Status")]},
        "notifications": {"key":"notifications", "label":"Notifications", "model":Notification, "columns":["id","title","message","notification_type","unread","created_at"], "fields":[("title","Title"),("message","Message"),("notification_type","Type"),("user_id","User ID"),("related_type","Related type"),("related_id","Related ID")]},
    }
    web_role_menus = {
        "Administrator": list(web_modules),
        "Super Administrator": list(web_modules),
        "Manager": ["products","customers","partners","stock","stock_movements","orders","sales","visits","plans","tasks","notifications","import_jobs"],
        "Sales/Agent": ["customers","partners","orders","sales","visits","tasks","notifications"],
        "Inventory Officer": ["products","inventory_locations","stock","stock_movements","requests","incidents","notifications"],
        "Viewer": ["products","customers","partners","stock","orders","sales","visits","tasks","notifications"],
        "ZBM": ["people","districts","territories","markets","booths","partners","resources","assignments","transfers","returns","tasks","audits","notifications"],
        "TSM": ["people","markets","booths","partners","resources","assignments","transfers","returns","requests","incidents","tasks","audits","notifications"],
        "TL": ["people","markets","booths","partners","resources","assignments","transfers","returns","requests","incidents","tasks","audits","notifications"],
        "TSE": ["markets","booths","partners","resources","requests","incidents","tasks","audits","notifications"],
        "Chabeba": ["markets","booths","partners","resources","tasks","incidents","notifications"],
    }

    def web_menu():
        keys = web_role_menus.get(session.get("role", "TSE"), web_role_menus["TSE"])
        items = [(web_modules[k]["label"], url_for("web_module", module_key=k)) for k in keys]
        items += [("Analytics", url_for("web_analytics")), ("Scheduler", url_for("web_scheduler")), ("Import Data", url_for("web_import")), ("Export Data", url_for("web_export")), ("Reports", url_for("web_reports")), ("Audit Logs", url_for("web_module", module_key="notifications")), ("Search", url_for("web_search"))]
        return items

    def web_required(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            if not session.get("user_id"):
                return redirect(url_for("web_login"))
            return fn(*args, **kwargs)
        return wrapper

    @app.get("/")
    def web_home():
        return redirect(url_for("web_dashboard") if session.get("user_id") else url_for("web_login"))

    @app.route("/login", methods=["GET", "POST"])
    def web_login():
        error = None
        if request.method == "POST":
            user = User.query.filter_by(username=request.form.get("username", "")).first()
            if user and user.check_password(request.form.get("password", "")):
                session["user_id"] = user.id
                session["username"] = user.username
                session["role"] = user.role
                log("web_login", user, "User logged in through web app"); db.session.commit()
                return redirect(url_for("web_dashboard"))
            error = "Invalid username or password"
        return render_template("login.html", error=error)

    @app.get("/logout")
    def web_logout():
        session.clear()
        return redirect(url_for("web_login"))

    @app.get("/dashboard")
    @web_required
    def web_dashboard():
        today = date.today()
        month_start = today.replace(day=1)
        total_stock = db.session.query(func.coalesce(func.sum(InventoryStock.quantity), 0)).scalar() or 0
        low_stock = db.session.query(Product).outerjoin(InventoryStock, InventoryStock.product_id == Product.id).group_by(Product.id).having(func.coalesce(func.sum(InventoryStock.quantity), 0) <= Product.minimum_stock).count()
        summary = {
            "total_products": Product.query.count(),
            "total_stock": total_stock,
            "low_stock_items": low_stock,
            "total_customers": Customer.query.count(),
            "total_partners": ChannelPartner.query.count(),
            "today_sales": Sale.query.filter_by(sale_date=today).count(),
            "monthly_sales": db.session.query(func.coalesce(func.sum(Sale.total_amount), 0)).filter(Sale.sale_date >= month_start).scalar() or 0,
            "pending_orders": Order.query.filter_by(status="Pending").count(),
            "completed_orders": Order.query.filter_by(status="Completed").count(),
            "upcoming_visits": Visit.query.filter(Visit.visit_date >= today, Visit.status.in_(["Scheduled", "Confirmed"])).count(),
            "pending_tasks": FieldTask.query.filter(FieldTask.status.in_(["Pending", "To Do", "In Progress", "Blocked"])).count(),
            "overdue_tasks": FieldTask.query.filter(FieldTask.due_date < today, FieldTask.status.notin_(["Completed", "Cancelled"])).count(),
            "upcoming_plan_reminders": Plan.query.filter(Plan.status == "Active", Plan.start_date >= today, Plan.start_date <= today + timedelta(days=7)).count(),
        }
        low_stock_rows = db.session.query(Product.sku, Product.name, Product.minimum_stock, func.coalesce(func.sum(InventoryStock.quantity), 0).label("quantity")).outerjoin(InventoryStock, InventoryStock.product_id == Product.id).group_by(Product.id).having(func.coalesce(func.sum(InventoryStock.quantity), 0) <= Product.minimum_stock).limit(10).all()
        upcoming_visits = Visit.query.filter(Visit.visit_date >= today).order_by(Visit.visit_date).limit(8).all()
        pending_tasks = FieldTask.query.filter(FieldTask.status.notin_(["Completed", "Cancelled"])).order_by(FieldTask.due_date).limit(8).all()
        return render_template("dashboard.html", summary=summary, low_stock_rows=low_stock_rows, upcoming_visits=upcoming_visits, pending_tasks=pending_tasks, menu=web_menu())

    @app.route("/modules/<module_key>")
    @web_required
    def web_module(module_key):
        module = web_modules.get(module_key)
        if not module: raise NotFound("Module not found")
        query = module["model"].query
        search = request.args.get("search")
        if search:
            cols = [c for c in module["model"].__table__.columns if str(c.type).startswith("VARCHAR")]
            if cols: query = query.filter(or_(*[c.ilike(f"%{search}%") for c in cols]))
        rows = [serialize(row) for row in query.limit(300).all()]
        return render_template("module.html", module=module, rows=rows, search=search, menu=web_menu())

    @app.route("/modules/<module_key>/new", methods=["GET", "POST"])
    @web_required
    def web_new_record(module_key):
        module = web_modules.get(module_key)
        if not module: raise NotFound("Module not found")
        error = None
        if request.method == "POST":
            try:
                payload = {k:v for k,v in request.form.items() if v}
                if module_key == "assignments":
                    item = ResourceAssignment(assigned_by_id=session.get("user_id"))
                    apply_payload(item, payload)
                    validate_resource_deployment(item.employee_id, item.market_id, item.booth_id, item.partner_id)
                    resource = Resource.query.get_or_404(item.resource_id)
                    resource.assigned_employee_id = item.employee_id
                    resource.market_id = item.market_id
                    resource.booth_id = item.booth_id
                    resource.partner_id = item.partner_id
                    resource.condition = item.condition
                    resource.status = "Assigned"
                elif module_key == "transfers":
                    item = ResourceTransfer(approved_by_id=session.get("user_id"))
                    apply_payload(item, payload)
                    booth_id = payload.get("booth_id")
                    partner_id = payload.get("partner_id")
                    validate_resource_deployment(item.to_employee_id, item.to_market_id, booth_id, partner_id)
                    resource = Resource.query.get_or_404(item.resource_id)
                    item.from_employee_id = resource.assigned_employee_id
                    item.from_market_id = resource.market_id
                    resource.assigned_employee_id = item.to_employee_id
                    resource.market_id = item.to_market_id
                    resource.booth_id = booth_id
                    resource.partner_id = partner_id
                    resource.status = "Transferred"
                elif module_key == "returns":
                    item = ResourceReturn(received_by_id=session.get("user_id"))
                    apply_payload(item, payload)
                    resource = Resource.query.get_or_404(item.resource_id)
                    resource.assigned_employee_id = None
                    resource.market_id = None
                    resource.booth_id = None
                    resource.partner_id = None
                    resource.status = "Returned"
                    resource.condition = item.condition or resource.condition
                else:
                    item = module["model"]()
                    apply_payload(item, payload)
                db.session.add(item); db.session.flush(); log("web_create", item); db.session.commit()
                return redirect(url_for("web_module", module_key=module_key))
            except Exception as exc:
                db.session.rollback(); error = str(exc)
        return render_template("form.html", module=module, error=error, menu=web_menu())

    @app.get("/search")
    @web_required
    def web_search():
        q_raw = request.args.get("q", "")
        results = {}
        if q_raw:
            q = f"%{q_raw}%"
            results = {
                "resources": [serialize(x) for x in Resource.query.filter(or_(Resource.resource_code.ilike(q), Resource.serial_number.ilike(q), Resource.imei.ilike(q), Resource.name.ilike(q))).limit(20)],
                "employees": [serialize(x) for x in Employee.query.filter(or_(Employee.employee_code.ilike(q), Employee.full_name.ilike(q))).limit(20)],
                "markets": [serialize(x) for x in Market.query.filter(or_(Market.market_code.ilike(q), Market.name.ilike(q))).limit(20)],
                "booths": [serialize(x) for x in Booth.query.filter(Booth.booth_code.ilike(q)).limit(20)],
                "partners": [serialize(x) for x in ChannelPartner.query.filter(or_(ChannelPartner.partner_code.ilike(q), ChannelPartner.business_name.ilike(q))).limit(20)],
            }
        return render_template("search.html", q=q_raw, results=results, menu=web_menu())

    @app.get("/reports")
    @web_required
    def web_reports():
        reports = ["products", "customers", "partners", "stock", "stock-movements", "orders", "sales", "visits", "plans", "resources", "assigned-resources", "lost-resources", "damaged-resources", "transfers", "returns", "markets", "booths", "channel-partners", "employees", "tasks", "audits"]
        return render_template("reports.html", reports=reports, menu=web_menu())

    @app.get("/reports/<report_name>.csv")
    @web_required
    def web_export_report(report_name):
        report_map = {
            "products": Product,
            "customers": Customer,
            "partners": ChannelPartner,
            "stock": InventoryStock,
            "stock-movements": StockMovement,
            "orders": Order,
            "sales": Sale,
            "visits": Visit,
            "plans": Plan,
            "resources": Resource,
            "assigned-resources": Resource,
            "lost-resources": Resource,
            "damaged-resources": Resource,
            "transfers": ResourceTransfer,
            "returns": ResourceReturn,
            "markets": Market,
            "booths": Booth,
            "channel-partners": ChannelPartner,
            "employees": Employee,
            "tasks": FieldTask,
            "audits": MarketAudit,
        }
        if report_name not in report_map:
            raise NotFound("Report not found")
        query = report_map[report_name].query
        if report_name == "assigned-resources":
            query = query.filter_by(status="Assigned")
        elif report_name == "lost-resources":
            query = query.filter_by(status="Lost")
        elif report_name == "damaged-resources":
            query = query.filter_by(status="Damaged")
        rows = [serialize(row) for row in query.limit(5000).all()]
        log("web_export_report", None, report_name); db.session.commit()
        return csv_response(f"{report_name}.csv", rows)

    IMPORT_FIELDS = {
        "Product": ["sku", "name", "category", "unit", "unit_price", "minimum_stock", "status"],
        "Customer": ["customer_code", "name", "phone", "email", "location", "status", "notes"],
        "Partner": ["partner_code", "business_name", "contact_person", "phone", "email", "market_id", "booth_id", "status", "notes"],
        "Inventory": ["sku", "location_code", "quantity"],
        "Stock Movement": ["movement_code", "sku", "from_location_code", "to_location_code", "quantity", "movement_type", "notes"],
        "Order": ["order_number", "customer_code", "partner_code", "status", "order_date", "due_date", "total_amount", "notes"],
        "Sale": ["sale_number", "customer_code", "partner_code", "sale_date", "total_amount", "status", "notes"],
        "Visit": ["title", "partner_code", "customer_code", "location", "visit_date", "start_time", "end_time", "purpose", "status", "priority", "notes"],
        "Task": ["title", "description", "assigned_employee_id", "due_date", "priority", "status", "notes"],
        "Plan": ["title", "description", "start_date", "end_date", "frequency", "assigned_employee_id", "priority", "status", "reminder_days"],
        "User": ["username", "role", "employee_id", "active"],
    }

    def import_field_guess(columns, fields):
        guesses = {}
        normalized = {str(c).strip().lower().replace(" ", "_").replace("-", "_"): c for c in columns}
        aliases = {"product_code": "sku", "product_name": "name", "customer_id": "customer_code", "partner_id": "partner_code", "business_name": "business_name", "qty": "quantity"}
        for field in fields:
            if field in normalized:
                guesses[normalized[field]] = field
        for source, target in aliases.items():
            if source in normalized and target in fields:
                guesses[normalized[source]] = target
        return guesses

    def workbook_response(filename, headers, example):
        wb = Workbook()
        ws = wb.active
        ws.title = "Template"
        ws.append(headers)
        ws.append(example)
        notes = wb.create_sheet("Instructions")
        notes.append(["Column", "Required", "Notes"])
        for h in headers:
            notes.append([h, "Yes" if h in headers[:2] else "No", "Map this column during import if your file uses a different header."])
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(tmp.name)
        return send_file(tmp.name, as_attachment=True, download_name=filename)

    @app.get("/import")
    @web_required
    def web_import():
        return render_template("import.html", data_types=IMPORT_FIELDS, menu=web_menu())

    @app.get("/export")
    @web_required
    def web_export():
        return render_template("export.html", reports=["products", "customers", "partners", "stock", "stock-movements", "orders", "sales", "visits", "tasks", "plans"], menu=web_menu())

    @app.get("/analytics")
    @web_required
    def web_analytics():
        return render_template("analytics.html", menu=web_menu())

    @app.get("/scheduler")
    @web_required
    def web_scheduler():
        visits = Visit.query.order_by(Visit.visit_date).limit(50).all()
        tasks = FieldTask.query.order_by(FieldTask.due_date).limit(50).all()
        plans = Plan.query.order_by(Plan.start_date).limit(50).all()
        return render_template("scheduler.html", visits=visits, tasks=tasks, plans=plans, menu=web_menu())

    @app.get("/templates/<data_type>.xlsx")
    @web_required
    def download_template(data_type):
        label = data_type.replace("-", " ").title()
        fields = IMPORT_FIELDS.get(label)
        if not fields:
            raise NotFound("Template not found")
        example = [f"Example {field}" if not field.endswith("_date") else date.today().isoformat() for field in fields]
        return workbook_response(f"marketlink-{data_type}-template.xlsx", fields, example)

    @app.post("/api/import/preview")
    @auth_required
    @roles_required("Administrator", "Super Administrator", "Manager")
    def import_preview():
        file = request.files.get("file")
        data_type = request.form.get("data_type", "Product")
        sheet_name = request.form.get("sheet")
        if not file or not file.filename.lower().endswith((".xlsx", ".xls")):
            raise BadRequest("Upload a valid Excel file")
        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(v).strip() for v in rows[0] if v is not None] if rows else []
        fields = IMPORT_FIELDS.get(data_type, IMPORT_FIELDS["Product"])
        mapping = import_field_guess(headers, fields)
        total = max(len(rows) - 1, 0)
        return jsonify(sheets=wb.sheetnames, selected_sheet=ws.title, data_type=data_type, total_rows=total, columns=headers, fields=fields, suggested_mapping=mapping)

    @app.post("/api/import/products")
    @auth_required
    @roles_required("Administrator", "Super Administrator", "Manager", "Inventory Officer")
    def import_products():
        file = request.files.get("file")
        duplicate_strategy = request.form.get("duplicate_strategy", "skip")
        if not file or not file.filename.lower().endswith((".xlsx", ".xls")):
            raise BadRequest("Upload a valid Excel file")
        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb[request.form.get("sheet")] if request.form.get("sheet") else wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise BadRequest("Workbook is empty")
        headers = [str(v).strip().lower().replace(" ", "_") for v in rows[0]]
        mapping = {h: h for h in headers}
        job = ImportJob(filename=secure_filename(file.filename), data_type="Product", sheet_name=ws.title, duplicate_strategy=duplicate_strategy, created_by_id=g.current_user.id, mapping=mapping)
        db.session.add(job); db.session.flush()
        imported = failed = skipped = 0
        for row_num, row in enumerate(rows[1:], start=2):
            data = {headers[i]: row[i] for i in range(min(len(headers), len(row))) if headers[i]}
            sku = str(data.get("sku") or data.get("product_code") or "").strip()
            name = str(data.get("name") or data.get("product_name") or "").strip()
            if not sku or not name:
                failed += 1
                db.session.add(ImportError(job_id=job.id, row_number=row_num, field="sku/name", problem="SKU and product name are required", original_value=str(row)))
                continue
            product = Product.query.filter_by(sku=sku).first()
            if product and duplicate_strategy == "skip":
                skipped += 1
                continue
            if not product:
                product = Product(sku=sku, name=name)
                db.session.add(product)
            product.name = name
            product.unit = str(data.get("unit") or product.unit or "Each")
            product.unit_price = float(data.get("unit_price") or data.get("price") or product.unit_price or 0)
            product.minimum_stock = int(data.get("minimum_stock") or data.get("min_stock") or product.minimum_stock or 0)
            product.status = str(data.get("status") or product.status or "Active")
            imported += 1
        job.total_rows = max(len(rows) - 1, 0)
        job.imported_rows = imported
        job.error_rows = failed
        job.valid_rows = imported
        job.warning_rows = skipped
        job.status = "Completed"
        log("excel_import", job, f"Imported products from {job.filename}")
        db.session.commit()
        return jsonify(job_id=job.id, total_rows=job.total_rows, imported=imported, skipped=skipped, failed=failed)

    @app.get("/health")
    def health():
        return jsonify(status="ok")

    @app.post("/api/auth/login")
    def login():
        data = request.get_json() or {}
        user = User.query.filter_by(username=data.get("username")).first()
        if not user or not user.check_password(data.get("password", "")):
            raise Unauthorized("Invalid username or password")
        token = user.issue_token(); log("login", user, "User logged in"); db.session.commit()
        return jsonify(token=token, user=serialize(user))

    @app.post("/api/auth/logout")
    @auth_required
    def logout():
        g.current_user.token = None; log("logout", g.current_user); db.session.commit()
        return jsonify(status="ok")

    for url, model in {
        "/api/users": User, "/api/employees": Employee, "/api/districts": District,
        "/api/territories": Territory, "/api/markets": Market, "/api/booths": Booth,
        "/api/channel-partners": ChannelPartner, "/api/resource-categories": ResourceCategory,
        "/api/resources": Resource, "/api/resource-requests": ResourceRequest,
        "/api/products": Product, "/api/customers": Customer, "/api/inventory-locations": InventoryLocation,
        "/api/stock": InventoryStock, "/api/stock-movements": StockMovement, "/api/orders": Order,
        "/api/sales": Sale, "/api/visits": Visit, "/api/plans": Plan,
        "/api/resource-incidents": ResourceIncident, "/api/tasks": FieldTask,
        "/api/audits": MarketAudit, "/api/notifications": Notification,
    }.items():
        crud_routes(url, model)

    @app.post("/api/resource-assignments")
    @auth_required
    @roles_required("Administrator", "ZBM", "TSM", "TL")
    def assign_resource():
        data = request.get_json() or {}
        assignment = ResourceAssignment(assigned_by_id=g.current_user.id)
        apply_payload(assignment, data)
        validate_resource_deployment(assignment.employee_id, assignment.market_id, assignment.booth_id, assignment.partner_id)
        resource = Resource.query.get_or_404(assignment.resource_id)
        resource.assigned_employee_id = assignment.employee_id
        resource.market_id = assignment.market_id
        resource.booth_id = assignment.booth_id
        resource.partner_id = assignment.partner_id
        resource.condition = assignment.condition
        resource.status = "Assigned"
        db.session.add(assignment); log("assign_resource", resource); db.session.commit()
        return jsonify(serialize(assignment)), 201

    @app.get("/api/resource-assignments")
    @auth_required
    def assignments():
        return jsonify([serialize(x) for x in ResourceAssignment.query.limit(200).all()])

    @app.post("/api/resource-transfers")
    @auth_required
    @roles_required("Administrator", "ZBM", "TSM", "TL")
    def transfer_resource():
        data = request.get_json() or {}
        transfer = ResourceTransfer(approved_by_id=g.current_user.id)
        apply_payload(transfer, data)
        booth_id = data.get("booth_id")
        partner_id = data.get("partner_id")
        validate_resource_deployment(transfer.to_employee_id, transfer.to_market_id, booth_id, partner_id)
        resource = Resource.query.get_or_404(transfer.resource_id)
        transfer.from_employee_id = resource.assigned_employee_id
        transfer.from_market_id = resource.market_id
        resource.assigned_employee_id = transfer.to_employee_id
        resource.market_id = transfer.to_market_id
        resource.booth_id = booth_id
        resource.partner_id = partner_id
        resource.status = "Transferred"
        db.session.add(transfer); log("transfer_resource", resource); db.session.commit()
        return jsonify(serialize(transfer)), 201

    @app.get("/api/resource-transfers")
    @auth_required
    def transfers():
        return jsonify([serialize(x) for x in ResourceTransfer.query.limit(200).all()])

    @app.post("/api/resource-returns")
    @auth_required
    @roles_required("Administrator", "ZBM", "TSM", "TL")
    def return_resource():
        data = request.get_json() or {}
        ret = ResourceReturn(received_by_id=g.current_user.id)
        apply_payload(ret, data)
        resource = Resource.query.get_or_404(ret.resource_id)
        resource.assigned_employee_id = None; resource.market_id = None; resource.booth_id = None; resource.partner_id = None
        resource.status = "Returned"; resource.condition = ret.condition or resource.condition
        db.session.add(ret); log("return_resource", resource); db.session.commit()
        return jsonify(serialize(ret)), 201

    @app.get("/api/resource-returns")
    @auth_required
    def returns():
        return jsonify([serialize(x) for x in ResourceReturn.query.limit(200).all()])

    @app.get("/api/search")
    @auth_required
    def search():
        q = f"%{request.args.get('q','')}%"
        resources = Resource.query.filter(or_(Resource.resource_code.ilike(q), Resource.serial_number.ilike(q), Resource.imei.ilike(q), Resource.name.ilike(q))).all()
        employees = Employee.query.filter(or_(Employee.employee_code.ilike(q), Employee.full_name.ilike(q))).all()
        markets = Market.query.filter(or_(Market.market_code.ilike(q), Market.name.ilike(q))).all()
        booths = Booth.query.filter(Booth.booth_code.ilike(q)).all()
        partners = ChannelPartner.query.filter(or_(ChannelPartner.partner_code.ilike(q), ChannelPartner.business_name.ilike(q))).all()
        return jsonify(resources=[serialize(x) for x in resources], employees=[serialize(x) for x in employees],
                       markets=[serialize(x) for x in markets], booths=[serialize(x) for x in booths],
                       partners=[serialize(x) for x in partners])

    @app.get("/api/resources/<int:item_id>/history")
    @auth_required
    def resource_history(item_id):
        Resource.query.get_or_404(item_id)
        return jsonify(assignments=[serialize(x) for x in ResourceAssignment.query.filter_by(resource_id=item_id)],
                       transfers=[serialize(x) for x in ResourceTransfer.query.filter_by(resource_id=item_id)],
                       returns=[serialize(x) for x in ResourceReturn.query.filter_by(resource_id=item_id)],
                       incidents=[serialize(x) for x in ResourceIncident.query.filter_by(resource_id=item_id)])


    @app.post("/api/uploads")
    @auth_required
    def upload_file():
        allowed = {"jpg", "jpeg", "png", "pdf", "docx", "xlsx"}
        file = request.files.get("file")
        if not file or not file.filename:
            raise BadRequest("A file is required")
        ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
        if ext not in allowed:
            raise BadRequest("This file type is not allowed")
        os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
        filename = secure_filename(file.filename)
        path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
        file.save(path)
        log("upload_file", None, filename); db.session.commit()
        return jsonify(filename=filename, path=path), 201
    @app.get("/api/reports/summary")
    @auth_required
    def report_summary():
        return jsonify(employees=Employee.query.count(), markets=Market.query.count(), booths=Booth.query.count(),
                       channel_partners=ChannelPartner.query.count(), resources=Resource.query.count(),
                       assigned=Resource.query.filter_by(status="Assigned").count(),
                       damaged=Resource.query.filter_by(status="Damaged").count(),
                       lost=Resource.query.filter_by(status="Lost").count(),
                       open_tasks=FieldTask.query.filter(FieldTask.status.in_(["To Do","In Progress","Blocked"])).count(),
                       unread_notifications=Notification.query.filter_by(user_id=g.current_user.id, unread=True).count())


    @app.get("/api/reports/<report_name>.csv")
    @auth_required
    def export_report(report_name):
        report_map = {
            "products": Product,
            "customers": Customer,
            "partners": ChannelPartner,
            "stock": InventoryStock,
            "stock-movements": StockMovement,
            "orders": Order,
            "sales": Sale,
            "visits": Visit,
            "plans": Plan,
            "resources": Resource,
            "assigned-resources": Resource,
            "lost-resources": Resource,
            "damaged-resources": Resource,
            "transfers": ResourceTransfer,
            "returns": ResourceReturn,
            "markets": Market,
            "booths": Booth,
            "channel-partners": ChannelPartner,
            "employees": Employee,
            "tasks": FieldTask,
            "audits": MarketAudit,
        }
        if report_name not in report_map:
            raise NotFound("Report not found")
        query = report_map[report_name].query
        if report_name == "assigned-resources":
            query = query.filter_by(status="Assigned")
        elif report_name == "lost-resources":
            query = query.filter_by(status="Lost")
        elif report_name == "damaged-resources":
            query = query.filter_by(status="Damaged")
        rows = [serialize(row) for row in query.limit(5000).all()]
        log("export_report", None, report_name); db.session.commit()
        return csv_response(f"{report_name}.csv", rows)
    @app.cli.command("db-info")
    def db_info():
        uri = app.config["SQLALCHEMY_DATABASE_URI"]
        safe_uri = uri
        if "://" in safe_uri and "@" in safe_uri:
            scheme, rest = safe_uri.split("://", 1)
            safe_uri = scheme + "://***:***@" + rest.split("@", 1)[1]
        print(f"Database URI: {safe_uri}")
        print(f"Provider: {'PostgreSQL/Neon' if uri.startswith('postgresql') else 'SQLite'}")
        print("Tables:")
        db.create_all()
        for table in sorted(db.metadata.tables):
            mapper = next((m for m in db.Model.registry.mappers if m.local_table.name == table), None)
            count = "?"
            if mapper:
                try:
                    count = db.session.query(mapper.class_).count()
                except Exception:
                    db.session.rollback()
            print(f"- {table}: {count}")

    @app.cli.command("export-db-csv")
    def export_db_csv():
        export_dir = os.path.join(tempfile.gettempdir(), "marketlink-export")
        os.makedirs(export_dir, exist_ok=True)
        db.create_all()
        exported = []
        for mapper in db.Model.registry.mappers:
            model = mapper.class_
            rows = [serialize(row) for row in model.query.limit(100000).all()]
            path = os.path.join(export_dir, f"{model.__tablename__}.csv")
            with open(path, "w", newline="", encoding="utf-8") as handle:
                if rows:
                    writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
                    writer.writeheader(); writer.writerows(rows)
                else:
                    handle.write("")
            exported.append(path)
        print("Exported CSV backup files:")
        for path in sorted(exported):
            print(path)

    @app.cli.command("reset-db")
    def reset_db():
        confirmation = os.getenv("MARKETLINK_RESET_CONFIRM", "")
        if confirmation != "RESET_AIRTEL_MARKETLINK":
            print("Refusing to reset database. Set MARKETLINK_RESET_CONFIRM=RESET_AIRTEL_MARKETLINK to proceed.")
            raise SystemExit(2)
        uri = app.config["SQLALCHEMY_DATABASE_URI"]
        print(f"Resetting Airtel MarketLink database provider: {'PostgreSQL/Neon' if uri.startswith('postgresql') else 'SQLite'}")
        db.drop_all()
        db.create_all()
        print("Database reset complete. Run `flask --app wsgi:app seed-data` only if you want seed/admin data.")

    @app.cli.command("init-db")
    def init_db():
        db.create_all(); print("Database initialized.")

    @app.cli.command("seed-data")
    def seed_data():
        seed(app)
        print("Demo data seeded. Login: admin / admin123")

    with app.app_context():
        db.create_all()
    return app

def seed(app):
    with app.app_context():
        existing_admin = User.query.filter_by(username="admin").first()
        if existing_admin:
            existing_admin.set_password("admin123")
            existing_admin.active = True
            db.session.commit()
            return
        cats = [ResourceCategory(name=n, group=g) for g, names in {
            "Devices": ["Smartphone", "Tablet", "Laptop", "POS", "MiFi", "Router", "Power bank"],
            "Clothing": ["Vendor jacket", "T-shirt", "Cap", "Reflective vest"],
            "Booth equipment": ["Table", "Chair", "Umbrella", "Display stand", "Signage"],
            "Marketing materials": ["Banner", "Poster", "Flyer", "Promotional material"],
        }.items() for n in names]
        db.session.add_all(cats); db.session.flush()
        d = District(district_code="DST-001", name="Lusaka", region="Central", status="Active")
        db.session.add(d); db.session.flush()
        zbm = Employee(employee_code="EMP-ZBM-001", full_name="Michael Banda", phone="0976000001", email="michael@example.com", position="ZBM", role="ZBM", district_id=d.id)
        db.session.add(zbm); db.session.flush(); d.zbm_id = zbm.id
        territories=[]; employees=[zbm]
        for i in range(2):
            tsm = Employee(employee_code=f"EMP-TSM-00{i+1}", full_name=f"TSM Demo {i+1}", role="TSM", district_id=d.id)
            db.session.add(tsm); db.session.flush()
            terr = Territory(territory_code=f"TER-00{i+1}", name=f"Lusaka {'Central' if i==0 else 'North'}", district_id=d.id, tsm_id=tsm.id)
            db.session.add(terr); db.session.flush(); tsm.territory_id = terr.id
            territories.append(terr); employees.append(tsm)
            for j in range(2):
                tl = Employee(employee_code=f"EMP-TL-{i+1}{j+1}", full_name=f"TL Demo {i+1}-{j+1}", role="TL", district_id=d.id, territory_id=terr.id, supervisor_id=tsm.id)
                db.session.add(tl); db.session.flush(); employees.append(tl)
                for k in range(3):
                    employees.append(Employee(employee_code=f"EMP-TSE-{i+1}{j+1}{k+1}", full_name=f"TSE Demo {i+1}-{j+1}-{k+1}", role="TSE", district_id=d.id, territory_id=terr.id, supervisor_id=tl.id))
                for k in range(4):
                    employees.append(Employee(employee_code=f"EMP-CHB-{i+1}{j+1}{k+1}", full_name=f"Chabeba Demo {i+1}-{j+1}-{k+1}", role="Chabeba", district_id=d.id, territory_id=terr.id, supervisor_id=tl.id))
        db.session.add_all(employees); db.session.flush()
        markets=[]
        for i in range(5):
            terr = territories[i % 2]
            m = Market(market_code=f"MKT-00{i+1}", name=["City Market","Matero Market","Chelstone Market","Kalingalinga Market","Chilenje Market"][i],
                       district_id=d.id, territory_id=terr.id, location="Lusaka", market_type="Urban", tsm_id=terr.tsm_id, tl_id=employees[3].id, tse_id=employees[5].id)
            db.session.add(m); db.session.flush(); markets.append(m)
        booths=[]
        for i in range(15):
            b = Booth(booth_code=f"BTH-{i+1:04d}", market_id=markets[i % 5].id, location=f"Section {chr(65+i%4)}", responsible_employee_id=employees[5].id)
            db.session.add(b); booths.append(b)
        db.session.flush()
        for i in range(30):
            db.session.add(ChannelPartner(partner_code=f"CP-{i+1:04d}", business_name=f"Demo Mobile Partner {i+1}", contact_person=f"Owner {i+1}",
                                          market_id=markets[i % 5].id, booth_id=booths[i % 15].id, tse_id=employees[5].id))
        for i in range(60):
            cat = cats[i % len(cats)]
            code = f"{cat.name[:3].upper().replace(' ','')}-{i+1:05d}"
            db.session.add(Resource(resource_code=code, name=f"{cat.name} {i+1}", brand="Demo Brand", category_id=cat.id, serial_number=f"SN{i+1:05d}", condition="Good"))
        admin = User(username="admin", role="Administrator", employee_id=zbm.id)
        admin.set_password("admin123")
        db.session.add(admin)
        db.session.add(FieldTask(title="Inspect all booth equipment", market_id=markets[0].id, booth_id=booths[0].id, assigned_employee_id=employees[3].id, priority="High", status="In Progress"))
        db.session.add(MarketAudit(market_id=markets[0].id, auditor_id=employees[3].id, checklist={"booth_exists": True, "branding_present": True}, discrepancies="Umbrella damaged"))
        db.session.commit()







